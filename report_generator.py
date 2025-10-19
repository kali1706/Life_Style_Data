from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.barcharts import VerticalBarChart
import pandas as pd
import openpyxl
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os
from datetime import datetime, timedelta
from models import User, WorkoutLog, NutritionLog, DailyStats
from utils import get_weekly_summary, get_macro_percentages, calculate_progress_score, get_fitness_recommendations
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import io
import base64

def generate_pdf_report(user_id):
    """Generate comprehensive PDF report for user"""
    user = User.query.get(user_id)
    if not user:
        raise ValueError("User not found")
    
    # Create reports directory if it doesn't exist
    os.makedirs('reports', exist_ok=True)
    
    filename = f'reports/lifestyle_report_{user.username}_{datetime.now().strftime("%Y%m%d")}.pdf'
    doc = SimpleDocTemplate(filename, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        textColor=colors.darkblue,
        alignment=1  # Center alignment
    )
    story.append(Paragraph(f"Lifestyle Analytics Report", title_style))
    story.append(Paragraph(f"User: {user.username}", styles['Heading2']))
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y')}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # User Profile Section
    story.append(Paragraph("User Profile", styles['Heading2']))
    profile_data = [
        ['Age', f"{user.age} years"],
        ['Gender', user.gender],
        ['Height', f"{user.height:.2f} m"],
        ['Weight', f"{user.weight:.1f} kg"],
        ['BMI', f"{user.bmi:.1f} ({user.get_bmi_category()})"],
        ['Experience Level', user.get_experience_level_text()],
        ['Body Fat %', f"{user.fat_percentage:.1f}%" if user.fat_percentage else "Not recorded"],
        ['Resting BPM', f"{user.resting_bpm} bpm" if user.resting_bpm else "Not recorded"]
    ]
    
    profile_table = Table(profile_data, colWidths=[2*inch, 3*inch])
    profile_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(profile_table)
    story.append(Spacer(1, 20))
    
    # Weekly Summary
    weekly_summary = get_weekly_summary(user_id)
    story.append(Paragraph("Weekly Summary (Last 7 Days)", styles['Heading2']))
    
    summary_data = [
        ['Metric', 'Value'],
        ['Total Workouts', str(weekly_summary['total_workouts'])],
        ['Total Calories Burned', f"{weekly_summary['total_calories_burned']:,} kcal"],
        ['Average Session Duration', f"{weekly_summary['avg_session_duration']:.1f} hours"],
        ['Total Meals Logged', str(weekly_summary['total_meals_logged'])],
        ['Average Daily Intake', f"{weekly_summary['avg_daily_intake']:.0f} kcal"],
        ['Average Water Intake', f"{weekly_summary['avg_water_intake']:.1f} L/day"],
        ['Consistency Score', f"{weekly_summary['consistency_score']:.1f}%"],
        ['Workout Types', ', '.join(weekly_summary['workout_types']) if weekly_summary['workout_types'] else 'None']
    ]
    
    summary_table = Table(summary_data, colWidths=[2.5*inch, 2.5*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # Macro Breakdown
    if weekly_summary['macro_percentages']['carbs'] > 0:
        story.append(Paragraph("Macronutrient Breakdown", styles['Heading2']))
        macro_data = [
            ['Macronutrient', 'Percentage', 'Ideal Range'],
            ['Carbohydrates', f"{weekly_summary['macro_percentages']['carbs']:.1f}%", "45-65%"],
            ['Proteins', f"{weekly_summary['macro_percentages']['proteins']:.1f}%", "15-25%"],
            ['Fats', f"{weekly_summary['macro_percentages']['fats']:.1f}%", "20-35%"]
        ]
        
        macro_table = Table(macro_data, colWidths=[2*inch, 1.5*inch, 1.5*inch])
        macro_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.green),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgreen),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(macro_table)
        story.append(Spacer(1, 20))
    
    # Progress Score
    progress_score = calculate_progress_score(user_id)
    story.append(Paragraph("Overall Progress Score", styles['Heading2']))
    score_color = colors.green if progress_score >= 80 else colors.orange if progress_score >= 60 else colors.red
    story.append(Paragraph(f"<font color='{score_color.hexval()}'><b>{progress_score}/100</b></font>", styles['Normal']))
    
    # Score interpretation
    if progress_score >= 80:
        interpretation = "Excellent! You're maintaining a very healthy lifestyle."
    elif progress_score >= 60:
        interpretation = "Good progress! Keep up the consistency and make small improvements."
    elif progress_score >= 40:
        interpretation = "Fair progress. Focus on improving workout consistency and nutrition balance."
    else:
        interpretation = "Needs improvement. Consider consulting with a fitness professional."
    
    story.append(Paragraph(interpretation, styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Recommendations
    recommendations = get_fitness_recommendations(user_id)
    if recommendations:
        story.append(Paragraph("Personalized Recommendations", styles['Heading2']))
        for i, rec in enumerate(recommendations, 1):
            priority_color = colors.red if rec['priority'] == 'high' else colors.orange if rec['priority'] == 'medium' else colors.blue
            story.append(Paragraph(f"{i}. <font color='{priority_color.hexval()}'>[{rec['priority'].upper()}]</font> {rec['message']}", styles['Normal']))
        story.append(Spacer(1, 20))
    
    # Recent Workout History
    recent_workouts = WorkoutLog.query.filter_by(user_id=user_id).order_by(WorkoutLog.date.desc()).limit(10).all()
    if recent_workouts:
        story.append(Paragraph("Recent Workout History", styles['Heading2']))
        workout_data = [['Date', 'Type', 'Duration (hrs)', 'Calories', 'Avg BPM']]
        for workout in recent_workouts:
            workout_data.append([
                workout.date.strftime('%Y-%m-%d'),
                workout.workout_type,
                f"{workout.session_duration:.1f}",
                str(workout.calories_burned),
                str(workout.avg_bpm) if workout.avg_bpm else 'N/A'
            ])
        
        workout_table = Table(workout_data, colWidths=[1*inch, 1.2*inch, 1*inch, 1*inch, 1*inch])
        workout_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.purple),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lavender),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(workout_table)
    
    # Build PDF
    doc.build(story)
    return filename

def generate_excel_report(user_id):
    """Generate comprehensive Excel report with multiple sheets"""
    user = User.query.get(user_id)
    if not user:
        raise ValueError("User not found")
    
    # Create reports directory if it doesn't exist
    os.makedirs('reports', exist_ok=True)
    
    filename = f'reports/lifestyle_data_{user.username}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    workbook = openpyxl.Workbook()
    
    # Remove default sheet
    workbook.remove(workbook.active)
    
    # Sheet 1: User Profile
    profile_sheet = workbook.create_sheet("User Profile")
    profile_sheet['A1'] = "User Profile Information"
    profile_sheet['A1'].font = Font(size=16, bold=True)
    profile_sheet['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    profile_sheet['A1'].font = Font(color="FFFFFF", size=16, bold=True)
    
    profile_data = [
        ["Field", "Value"],
        ["Username", user.username],
        ["Email", user.email],
        ["Age", f"{user.age} years"],
        ["Gender", user.gender],
        ["Height", f"{user.height:.2f} m"],
        ["Weight", f"{user.weight:.1f} kg"],
        ["BMI", f"{user.bmi:.1f}"],
        ["BMI Category", user.get_bmi_category()],
        ["Experience Level", user.get_experience_level_text()],
        ["Body Fat %", f"{user.fat_percentage:.1f}%" if user.fat_percentage else "Not recorded"],
        ["Resting BPM", f"{user.resting_bpm} bpm" if user.resting_bpm else "Not recorded"]
    ]
    
    for row_num, row_data in enumerate(profile_data, 3):
        for col_num, cell_data in enumerate(row_data, 1):
            cell = profile_sheet.cell(row=row_num, column=col_num, value=cell_data)
            if row_num == 3:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Sheet 2: Workout History
    workout_sheet = workbook.create_sheet("Workout History")
    workouts = WorkoutLog.query.filter_by(user_id=user_id).order_by(WorkoutLog.date.desc()).all()
    
    workout_sheet['A1'] = "Workout History"
    workout_sheet['A1'].font = Font(size=16, bold=True)
    workout_sheet['A1'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    workout_sheet['A1'].font = Font(color="FFFFFF", size=16, bold=True)
    
    workout_headers = ["Date", "Workout Type", "Duration (hrs)", "Calories Burned", "Max BPM", "Avg BPM", "Intensity"]
    for col_num, header in enumerate(workout_headers, 1):
        cell = workout_sheet.cell(row=3, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    for row_num, workout in enumerate(workouts, 4):
        workout_sheet.cell(row=row_num, column=1, value=workout.date.strftime('%Y-%m-%d'))
        workout_sheet.cell(row=row_num, column=2, value=workout.workout_type)
        workout_sheet.cell(row=row_num, column=3, value=workout.session_duration)
        workout_sheet.cell(row=row_num, column=4, value=workout.calories_burned)
        workout_sheet.cell(row=row_num, column=5, value=workout.max_bpm or 0)
        workout_sheet.cell(row=row_num, column=6, value=workout.avg_bpm or 0)
        workout_sheet.cell(row=row_num, column=7, value=workout.intensity_level or "N/A")
    
    # Sheet 3: Nutrition History
    nutrition_sheet = workbook.create_sheet("Nutrition History")
    nutrition_logs = NutritionLog.query.filter_by(user_id=user_id).order_by(NutritionLog.date.desc()).all()
    
    nutrition_sheet['A1'] = "Nutrition History"
    nutrition_sheet['A1'].font = Font(size=16, bold=True)
    nutrition_sheet['A1'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    nutrition_sheet['A1'].font = Font(color="FFFFFF", size=16, bold=True)
    
    nutrition_headers = ["Date", "Meal Name", "Calories", "Carbs (g)", "Proteins (g)", "Fats (g)", "Water (L)", "Is Healthy"]
    for col_num, header in enumerate(nutrition_headers, 1):
        cell = nutrition_sheet.cell(row=3, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    for row_num, nutrition in enumerate(nutrition_logs, 4):
        nutrition_sheet.cell(row=row_num, column=1, value=nutrition.date.strftime('%Y-%m-%d'))
        nutrition_sheet.cell(row=row_num, column=2, value=nutrition.meal_name)
        nutrition_sheet.cell(row=row_num, column=3, value=nutrition.calories)
        nutrition_sheet.cell(row=row_num, column=4, value=nutrition.carbs)
        nutrition_sheet.cell(row=row_num, column=5, value=nutrition.proteins)
        nutrition_sheet.cell(row=row_num, column=6, value=nutrition.fats)
        nutrition_sheet.cell(row=row_num, column=7, value=nutrition.water_intake)
        nutrition_sheet.cell(row=row_num, column=8, value="Yes" if nutrition.is_healthy else "No")
    
    # Sheet 4: Weekly Summary
    summary_sheet = workbook.create_sheet("Weekly Summary")
    weekly_summary = get_weekly_summary(user_id)
    
    summary_sheet['A1'] = "Weekly Summary (Last 7 Days)"
    summary_sheet['A1'].font = Font(size=16, bold=True)
    summary_sheet['A1'].fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
    summary_sheet['A1'].font = Font(color="FFFFFF", size=16, bold=True)
    
    summary_data = [
        ["Metric", "Value"],
        ["Total Workouts", weekly_summary['total_workouts']],
        ["Total Calories Burned", f"{weekly_summary['total_calories_burned']:,} kcal"],
        ["Average Session Duration", f"{weekly_summary['avg_session_duration']:.1f} hours"],
        ["Total Meals Logged", weekly_summary['total_meals_logged']],
        ["Average Daily Intake", f"{weekly_summary['avg_daily_intake']:.0f} kcal"],
        ["Average Water Intake", f"{weekly_summary['avg_water_intake']:.1f} L/day"],
        ["Consistency Score", f"{weekly_summary['consistency_score']:.1f}%"],
        ["Progress Score", f"{calculate_progress_score(user_id)}/100"],
        ["Carbs %", f"{weekly_summary['macro_percentages']['carbs']:.1f}%"],
        ["Proteins %", f"{weekly_summary['macro_percentages']['proteins']:.1f}%"],
        ["Fats %", f"{weekly_summary['macro_percentages']['fats']:.1f}%"]
    ]
    
    for row_num, row_data in enumerate(summary_data, 3):
        for col_num, cell_data in enumerate(row_data, 1):
            cell = summary_sheet.cell(row=row_num, column=col_num, value=cell_data)
            if row_num == 3:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    
    # Sheet 5: Monthly Trends (if enough data)
    thirty_days_ago = datetime.now() - timedelta(days=30)
    monthly_workouts = WorkoutLog.query.filter(
        WorkoutLog.user_id == user_id,
        WorkoutLog.date >= thirty_days_ago.date()
    ).order_by(WorkoutLog.date).all()
    
    if monthly_workouts:
        trends_sheet = workbook.create_sheet("Monthly Trends")
        trends_sheet['A1'] = "Monthly Trends (Last 30 Days)"
        trends_sheet['A1'].font = Font(size=16, bold=True)
        trends_sheet['A1'].fill = PatternFill(start_color="9966CC", end_color="9966CC", fill_type="solid")
        trends_sheet['A1'].font = Font(color="FFFFFF", size=16, bold=True)
        
        # Group workouts by date
        daily_data = {}
        for workout in monthly_workouts:
            date_str = workout.date.strftime('%Y-%m-%d')
            if date_str not in daily_data:
                daily_data[date_str] = {'calories': 0, 'duration': 0, 'count': 0}
            daily_data[date_str]['calories'] += workout.calories_burned
            daily_data[date_str]['duration'] += workout.session_duration
            daily_data[date_str]['count'] += 1
        
        trends_headers = ["Date", "Total Calories Burned", "Total Duration (hrs)", "Workout Count"]
        for col_num, header in enumerate(trends_headers, 1):
            cell = trends_sheet.cell(row=3, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E1D5E7", end_color="E1D5E7", fill_type="solid")
        
        for row_num, (date_str, data) in enumerate(sorted(daily_data.items()), 4):
            trends_sheet.cell(row=row_num, column=1, value=date_str)
            trends_sheet.cell(row=row_num, column=2, value=data['calories'])
            trends_sheet.cell(row=row_num, column=3, value=data['duration'])
            trends_sheet.cell(row=row_num, column=4, value=data['count'])
    
    # Auto-adjust column widths
    for sheet in workbook.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(filename)
    return filename

def send_email_report(user_id, recipient_email, report_type='pdf'):
    """Send email report to specified email address"""
    user = User.query.get(user_id)
    if not user:
        raise ValueError("User not found")
    
    # Generate report
    if report_type == 'pdf':
        attachment_file = generate_pdf_report(user_id)
        attachment_name = f'lifestyle_report_{user.username}.pdf'
    else:
        attachment_file = generate_excel_report(user_id)
        attachment_name = f'lifestyle_data_{user.username}.xlsx'
    
    # Email configuration (you'll need to set these in your environment)
    smtp_server = os.environ.get('MAIL_SERVER', 'smtp.gmail.com')
    smtp_port = int(os.environ.get('MAIL_PORT', 587))
    sender_email = os.environ.get('MAIL_USERNAME')
    sender_password = os.environ.get('MAIL_PASSWORD')
    
    if not sender_email or not sender_password:
        raise ValueError("Email configuration not set. Please configure MAIL_USERNAME and MAIL_PASSWORD environment variables.")
    
    # Create message
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = recipient_email
    msg['Subject'] = f"Your Lifestyle Analytics Report - {datetime.now().strftime('%B %d, %Y')}"
    
    # Get weekly summary for email body
    weekly_summary = get_weekly_summary(user_id)
    progress_score = calculate_progress_score(user_id)
    
    body = f"""
    Dear {user.username},

    Here's your detailed lifestyle analytics report for the week ending {datetime.now().strftime('%B %d, %Y')}.

    📊 WEEKLY HIGHLIGHTS:
    • Workouts Completed: {weekly_summary['total_workouts']}
    • Total Calories Burned: {weekly_summary['total_calories_burned']:,} kcal
    • Average Session Duration: {weekly_summary['avg_session_duration']:.1f} hours
    • Average Daily Intake: {weekly_summary['avg_daily_intake']:.0f} kcal
    • Consistency Score: {weekly_summary['consistency_score']:.1f}%
    • Overall Progress Score: {progress_score}/100

    🎯 MACRO BREAKDOWN:
    • Carbohydrates: {weekly_summary['macro_percentages']['carbs']:.1f}%
    • Proteins: {weekly_summary['macro_percentages']['proteins']:.1f}%
    • Fats: {weekly_summary['macro_percentages']['fats']:.1f}%

    📈 CURRENT STATS:
    • BMI: {user.bmi:.1f} ({user.get_bmi_category()})
    • Weight: {user.weight:.1f} kg
    • Experience Level: {user.get_experience_level_text()}

    Your detailed report is attached to this email with comprehensive charts, trends, and personalized recommendations.

    Keep up the great work on your fitness journey!

    Best regards,
    Lifestyle Analytics Team

    ---
    This is an automated report. Please do not reply to this email.
    """
    
    msg.attach(MIMEText(body, 'plain'))
    
    # Attach report file
    with open(attachment_file, "rb") as attachment:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment.read())
    
    encoders.encode_base64(part)
    part.add_header(
        'Content-Disposition',
        f'attachment; filename= {attachment_name}',
    )
    msg.attach(part)
    
    # Send email
    server = smtplib.SMTP(smtp_server, smtp_port)
    server.starttls()
    server.login(sender_email, sender_password)
    text = msg.as_string()
    server.sendmail(sender_email, recipient_email, text)
    server.quit()
    
    # Clean up temporary file
    if os.path.exists(attachment_file):
        os.remove(attachment_file)
    
    return True

def create_chart_image(chart_type, data, title="Chart"):
    """Create chart image for reports"""
    plt.figure(figsize=(10, 6))
    
    if chart_type == 'pie':
        plt.pie(data['values'], labels=data['labels'], autopct='%1.1f%%', startangle=90)
        plt.title(title)
    elif chart_type == 'bar':
        plt.bar(data['labels'], data['values'])
        plt.title(title)
        plt.xticks(rotation=45)
    elif chart_type == 'line':
        plt.plot(data['x'], data['y'])
        plt.title(title)
        plt.xticks(rotation=45)
    
    # Save to bytes
    img_buffer = io.BytesIO()
    plt.savefig(img_buffer, format='png', bbox_inches='tight', dpi=150)
    img_buffer.seek(0)
    plt.close()
    
    return img_buffer